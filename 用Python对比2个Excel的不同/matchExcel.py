#!/usr/bin/env python3
from openpyxl.styles import PatternFill
from openpyxl.styles import colors
from openpyxl.styles import Font
import openpyxl as pxl

"""
比较2个excel的数据不同地方,并且用黄色标记出来.
"""

workbook1 = pxl.load_workbook(r"/Users/luochuan/Desktop/用Python对比2个Excel的不同/Resume_Louis.Luo_Test1.xlsx")
workbook2 = pxl.load_workbook(r"/Users/luochuan/Desktop/用Python对比2个Excel的不同/Resume_Louis.Luo_test2.xlsx")


workbook1_sheet1 = workbook1['sheet']
workbook2_sheet1 = workbook2['sheet']


if workbook1_sheet1.max_row > workbook2_sheet1.max_row:
    max_row = workbook1_sheet1.max_row
else:
    max_row = workbook2_sheet1.max_row
    
if workbook1_sheet1.max_column > workbook2_sheet1.max_column:
    max_column = workbook1_sheet1.max_column
else:
    max_column = workbook2_sheet1.max_column
    
    
    
for i in range(1, (max_row + 1)):
    for j in range(1, (max_column +1)):
        cell_1 = workbook1_sheet1.cell(i, j)
        cell_2 = workbook2_sheet1.cell(i, j)
        
        if cell_1.value != cell_2.value:
            cell_1.fill = PatternFill("solid", fgColor='FFFF00')
            cell_1.font = Font(color=colors.BLACK, bold=True)
            cell_2.fill = PatternFill("solid", fgColor='FFFF00')
            cell_2.font = Font(color=colors.BLACK, bold=True)

workbook1.save(r'result1.xlsx')
workbook2.save(r'result2.xlsx')
