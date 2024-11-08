#!/usr/bin/env python3

import openpyxl as pxl
import csv 
import os

path = os.getcwd()
excelFileName = path + "/rules.xlsx"
csvFileName = path + "/testData.csv"

workBook1 = pxl.load_workbook(excelFileName)
workBook1_sheet = workBook1['Sheet1']

# 获取映射表数据(rules.xlsx)
ruleList = []
for i in range(1, (workBook1_sheet.max_row +1)):
    nameStr = workBook1_sheet.cell(i, 1).value
    ageStr = workBook1_sheet.cell(i, 2).value
    idStr = workBook1_sheet.cell(i, 3).value
    if len(idStr) > 0:
        ruleOneList = [nameStr,ageStr,idStr]
        ruleList.append(ruleOneList)
  
# 读取源数据(testData.csv) 并且根据映射表数据规则 修改源数据   
rows = []
with open(csvFileName, 'r', newline='', encoding='utf-8') as f:
    rows = [row for row in csv.DictReader(f)]
    for i in range(0, len(rows)):
        if rows[i]['Id'] == "Null":
            for j in range(0, len(ruleList)):
                if rows[i]['Name'] == str(ruleList[j][0]) and rows[i]['Age'] == str(ruleList[j][1]) :
                    rows[i]['Id'] = str(ruleList[j][2])
                                                             
f.close
                 
# 将修改后的源数据写入源数据文件中           
header = rows[0].keys()  
with open(csvFileName, 'w', newline='', encoding='utf-8') as f:
    writer = csv.DictWriter(f, fieldnames=header)
    writer.writeheader()
    writer.writerows(rows)
       
f.close
                              







