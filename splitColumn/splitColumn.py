#!/usr/bin/env python3

from openpyxl.styles import PatternFill 
from openpyxl.styles import colors
from openpyxl.styles import Font
from openpyxl.styles import Border
from openpyxl.styles import Side
import openpyxl as pxl

"""
将excel某一列中数据拆分为多行, 例如第三列中的值按照空格分割成多行
"""

workbook1 = pxl.load_workbook(r"/Users/luochuan/Desktop/test.xlsx")
workbook1_sheet1 = workbook1['sheet1']

splitString = "\n"


for i in range(1, (workbook1_sheet1.max_row)):
    value = workbook1_sheet1.cell(i, 3).value
    if value != None:
        list = value.split(splitString)
        if len(list) > 1:
            
            workbook1_sheet1.cell(i, 3).value = list[0]
            workbook1_sheet1.insert_rows(idx=(i+1), amount= (len(list)-1))


            for j in range(1, len(list)):
                
                for col in range(1, workbook1_sheet1.max_column+1):
                    cell = workbook1_sheet1.cell(i, column = col)
                    targetCell = workbook1_sheet1.cell(row=(i+j), column=col)
                    if col == 3: 
                        
                        targetCell.value = list[j]
                    else:
                        
                        targetCell.value = cell.value
                    
                    # targetCell.fill=PatternFill("solid", fgColor='FFFF00')
                    targetCell.font = Font(color=colors.BLACK, bold=False, size=16)
                    targetCell.border = Border(left=Side(border_style='double', color='000000'),
                right=Side(border_style='double', color='000000'),
                top=Side(border_style='double', color='000000'),
                bottom=Side(border_style='double', color='000000'))
                    
workbook1.save(r'test11111.xlsx')